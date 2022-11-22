from openpyxl import load_workbook
import splitter
from analyzeComments import reduceComments

dataFile = 'ExtractFunctions'
dataFile = dataFile + '.xlsx'

wbData = load_workbook(filename = dataFile)

# grab the active worksheet
wsData = wbData.active
row_countData = wsData.max_row

keyWordFile = 'keyWords.xlsx'
wbKeyWord = load_workbook(filename = keyWordFile)

# grab the active worksheet
wsKeyWords = wbKeyWord.active
row_count = wsKeyWords.max_row

hitsArray = []
groupVal = ''
colData = 4
for row in range(2, row_countData):
    getCell = wsData.cell(row, colData)
    funcName = getCell.value
    funcNameSplit = funcName.split('_')
    # print(funcNameSplit)

    # analyse comments
    getComments = reduceComments(row, colData, wsData)
    
    for row1 in wsKeyWords.iter_cols(min_row = 2, min_col =1, max_col=7, max_row=8):
        for cell in row1:
            if not (cell.value == None): 
                # print(cell.row, ' ', cell.column)
                # print(cell.value)
                # funcNameSplit = (splitter.split(funcName))
                # funcNameSplit = funcName.split('_')
                # print(funcNameSplit)
                for eachWord in funcNameSplit:
                    if cell.value in eachWord.lower() : #in funcName: #check if keywords are in function name
                        hitsArray.append(funcName + ' ' + wsKeyWords.cell(1, cell.column).value) #if yes, save top level classification
                        groupVal = wsKeyWords.cell(1, cell.column).value
                        if wsData.cell(row,colData+4+ cell.column).value == None:
                            wsData.cell(row,colData+4+ cell.column).value =  1  #increment vote
                        else:
                            wsData.cell(row,colData+4+ cell.column).value = wsData.cell(row,colData+4+ cell.column).value + 1
                        print(funcName + ' ---> ' + groupVal)
                if not getComments == None :
                    for eachword in getComments:
                        if cell.value in eachWord.lower() : #in funcName: #check if keywords are in function name
                            if wsData.cell(row,colData+4+ cell.column).value == None:
                                wsData.cell(row,colData+4+ cell.column).value =  1  #increment vote
                            else:
                                wsData.cell(row,colData+4+ cell.column).value = wsData.cell(row,colData+4+ cell.column).value + 1
                                
    hitsArrayVal = ' '.join([''.join(sub) for sub in hitsArray])
    wsData.cell(row, colData+3).value = hitsArrayVal
    wsData.cell(row, + colData+4).value = groupVal
    # print(hitsArray)
    hitsArray=[] # re-init array for next function name search 
    groupVal = '' # re-init for next row

wsData.cell(1, colData+3).value = 'Function+Classification'
wsData.cell(1, + colData+4).value = 'Classification'

# classification columns 
wsData.cell(1, + colData+5).value = 'Sensor Functions'
wsData.cell(1, + colData+6).value = 'Control Functions'
wsData.cell(1, + colData+7).value = 'Actuator Functions'
wsData.cell(1, + colData+8).value = 'Logging and Diagnostics'
wsData.cell(1, + colData+9).value = 'Initialization'
wsData.cell(1, + colData+10).value = 'Calibration'
wsData.cell(1, + colData+11).value = 'Communication'

wbData.save(filename = dataFile)
# print(len(hitsArray))

