from openpyxl import load_workbook

#dataFile = 'ExtractFunctions1.xlsx'
def getComments(row, column, ws):
    #row = 633
    #column = 1
   
    getCell = ws.cell(row, column) #633
    cellVal = getCell.value
    print(cellVal)
    stringSplit = cellVal.split('\\')
    folderName = stringSplit[0]
    stringSplit = stringSplit[1].split('.cpp_')
    cppFileName = stringSplit[0]+'.cpp'
    funcName = stringSplit[1]
    print(cppFileName)
    # print(funcName)
    ws.cell(row, column+2).value = funcName

    getCell = ws.cell(row, column+1)
    cellVal = getCell.value
    funcName = (cellVal.partition('\n')[0])
    print(funcName)

    cppFilePath = folderName + '\\' + cppFileName 

    list_of_lines = []
    with open(cppFilePath, 'r') as fp:
        for l_no, line in enumerate(fp):
            # search string
            if funcName in line:
                # print('string found in a file')
                # print('Line Number:', l_no)
                print('Line:', line)
                # don't look for next lines
                list_of_lines.append((l_no, line.rstrip()))
    fp.close()
    # print(len(list_of_lines))
    # print(list_of_lines)
    # for elem in list_of_lines:
    #         print(elem[0])

    readFile = open(cppFilePath)
    # read the content of the file opened
    fileContent = readFile.readlines()

    arrComments = []
    stopLim = ['}' , ';' , '#']
    for elem in list_of_lines:
        l_no = elem[0]
        for num in reversed(range(l_no)):
            # print(num)
            line = fileContent[num]
            # line = line.rstrip()
            # print(line)
            if any(x in line for x in stopLim):
                # print(l_no)
                break
            else:
                # print(l_no)
                t_line = str(line).rstrip('\n')
                arrComments.append(t_line)

    fileContent = []
    # print(arrComments)
    # Using list comprehension + join()
    arrCommentsVal = ' '.join([''.join(sub) for sub in arrComments])
    print(arrCommentsVal)
    ws.cell(row, column+3).value = arrCommentsVal

    

