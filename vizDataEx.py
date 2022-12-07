#Import the required library
from tkinter import*
from openpyxl import load_workbook

#Create an instance of tkinter frame
# win= Tk()
#Set the geometry
# win.geometry("1600x1200")

# dataFile = 'ExtractFunctions1'
# dataFile = dataFile + '.xlsx'

def createViz(win, dataFile):
    #Create an canvas object
    canvas= Canvas(win, width= 800, height= 880, bg = "white", relief="ridge")


    wbData = load_workbook(filename = dataFile)

    # grab the active worksheet
    wsData = wbData.active
    row_countData = wsData.max_row

    colData = 3
    sensorFcn = []
    cntrlFcn = []
    actrFcn = []
    logFcn = []
    initFcn = []
    calFcn = []
    commFcn = []
    noClassFcn = []

    for row in range(2, row_countData):
        getClass = wsData.cell(row, colData+14).value
        fcnANDcode = [wsData.cell(row,colData+1).value , wsData.cell(row,colData).value]
        if getClass == wsData.cell(1, colData+6).value:
            sensorFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+7).value:
            cntrlFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+8).value:
            actrFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+9).value:
            logFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+10).value:
            initFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+11).value:
            calFcn.append(fcnANDcode)
        elif getClass == wsData.cell(1, colData+12).value:
            commFcn.append(fcnANDcode)
        else:
            noClassFcn.append(fcnANDcode)


    # print(len(sensorFcn), ' ' , len(cntrlFcn), ' ', len(actrFcn), ' ' , len(logFcn), ' ', len(initFcn), ' ' , len(noClassFcn))
    # arrayClass = [[sensorFcn], [cntrlFcn], [actrFcn], [logFcn], [initFcn], [commFcn], [noClassFcn]]
    arrayClass = []
    arrayClass.append(sensorFcn)
    arrayClass.append(cntrlFcn)
    arrayClass.append(actrFcn)
    arrayClass.append(logFcn)
    arrayClass.append(initFcn)
    arrayClass.append(calFcn)
    arrayClass.append(commFcn)
    arrayClass.append(noClassFcn)
    # print(len(arrayClass)) #6


    def createDot(item, xshift, listDots, colorSelect):
        j = 0
        k = 0
        m = 0
        # getCell = [wsData.cell(row=4, column=4).value, wsData.cell(row=4, column=3).value]
        for i in range(0,len(item)):
            gettags1 = (item[i])
            # print(gettags1)
            x0 = 10 + j + xshift
            y0 = 10 + k
            x1 = 20 + j + xshift
            y1 = 20 + k

            m = m + 1

            if not m % 4:
                j = 0
            else:
                j = j + 15
            
            if m % 20:
                k = k + 5

            # k = k + 5
            # j = j + 5
            c  = (canvas.create_oval(x0,y0,x1,y1, outline=colorSelect, fill=colorSelect, tag=gettags1))
            listDots.append(c)
            # print(i, ' ', x0, ' ', y0, ' ', x1, ' ', y1)

    colorSelect = ['grey','red','green','blue','cyan','yellow','magenta','purple']

    listDots = []
    xshift = 0
    k=0

    # print((arrayClass[0]))



    for item in arrayClass:
        createDot(item, xshift, listDots, colorSelect[k])
        xshift = xshift + 100
        # print(print(item[1]))
        k = k+1
        




    # print(getCell[1])
    # print(sensorFcn[2][1])
    # print(sensorFcn[82])
    # print(len(sensorFcn))

    # A = canvas.create_oval(20,20,30,30, outline='grey', fill="grey", tag=getCell)
    # B = canvas.create_oval(130,130,140,140, outline='green', fill="green", tag=["B","second"])

    lbl = Label(win)
    lbl.place(x=820, y=20)

    lbl2 = Label(win, justify=LEFT)
    lbl2.place(x=820, y=40)

    k = 0
    for col in range(9,17):
        lblClass = Label(win)
        lblTxt = wsData.cell(1, col).value
        # print(lblTxt)
        if lblTxt == None:
            lblTxt = 'Unclassified'
        lblClass.config(text=lblTxt)
        lblClass.place(x=20 + k,y=100)
        k = k+100

    def on_enter(e):
        # find the canvas item below mouse cursor
        item = canvas.find_withtag("current")
        # get the tags for the item
        tags = canvas.gettags(item)
        # show it using the label
        lbl.config(text=tags[0])

        def displayTag(e):
            lbl2.config(text = tags[1])

        canvas.tag_bind(item, "<Button-1>", displayTag)

    def on_leave(e):
        # clear the label text
        lbl.config(text="")

    for item in (listDots):
        canvas.tag_bind(item, "<Enter>", on_enter)
        canvas.tag_bind(item, "<Leave>", on_leave)

    # canvas.pack()
    # canvas.place(relx=0.3, rely=1, anchor=S)
    canvas.place(x=420, y=1000, anchor=S)
    # win.mainloop()
    wbData.close()