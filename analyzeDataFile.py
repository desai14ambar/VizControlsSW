from openpyxl import load_workbook
from analyzeComments import reduceComments

# dataFile = 'ExtractFunctions1'
# dataFile = dataFile + '.xlsx'

def analyzeData(dataFile):

    wbData = load_workbook(filename = dataFile)

    # grab the active worksheet
    wsData = wbData.active
    row_countData = wsData.max_row

    keyWordFile = 'keyWords.xlsx'
    wbKeyWord = load_workbook(filename = keyWordFile)

    # grab the active worksheet
    wsKeyWords = wbKeyWord.active
    row_countKeyWords = wsKeyWords.max_row

    hitsArray = []
    groupVal = ''
    colData = 4
    for row in range(2, row_countData):
        getCell = wsData.cell(row, colData)
        funcName = getCell.value
        funcNameSplit = funcName.split('_')
        # print(funcNameSplit)
        
        for row1 in wsKeyWords.iter_cols(min_row = 2, min_col =1, max_col=7, max_row=row_countKeyWords):
            for cell in row1:
                if not (cell.value == None): 
                    # print(cell.row, ' ', cell.column)
                    # print(cell.value)
                    # funcNameSplit = (splitter.split(funcName))
                    # funcNameSplit = funcName.split('_')
                    # print(funcNameSplit)
                    for eachWord in funcNameSplit:
                        if cell.value in eachWord.lower() : #in funcName: #check if keywords are in function name
                            hitsArray.append(funcName + ' ' + wsKeyWords.cell(1, cell.column).value) #if yes, save top level classification
                            groupVal = wsKeyWords.cell(1, cell.column).value
                            if wsData.cell(row,colData+4+ cell.column).value == None:
                                wsData.cell(row,colData+4+ cell.column).value =  1.1  #increment vote
                            else:
                                wsData.cell(row,colData+4+ cell.column).value = wsData.cell(row,colData+4+ cell.column).value + 1.1
                            print(funcName + ' ---> ' + groupVal)
                    # if not getComments == None :
                    #     for eachComment in getComments:
                    #         if cell.value in eachComment.lower() : #in funcName: #check if keywords are in function name
                    #             print(cell.value + ' ---> ' + eachComment)
                    #             if wsData.cell(row,colData+4+ cell.column).value == None:
                    #                 wsData.cell(row,colData+4+ cell.column).value =  1  #increment vote
                    #             else:
                    #                 wsData.cell(row,colData+4+ cell.column).value = wsData.cell(row,colData+4+ cell.column).value + 1

        hitsArrayVal = ' '.join([''.join(sub) for sub in hitsArray])
        wsData.cell(row, colData+3).value = hitsArrayVal
        wsData.cell(row, + colData+4).value = groupVal
        # print(hitsArray)
        hitsArray=[] # re-init array for next function name search 
        groupVal = '' # re-init for next row

        # analyse comments
        getComments = reduceComments(row, colData, wsData)
        if not getComments == None:
            for row1 in wsKeyWords.iter_cols(min_row = 2, min_col =1, max_col=7, max_row=row_countKeyWords):
                for cell in row1:
                    if not (cell.value == None): 
                        for eachComment in getComments:
                            if cell.value in eachComment.lower() : #in funcName: #check if keywords are in function name
                                print(cell.value + ' ---> ' + eachComment + '  -  COMMENT')
                                if wsData.cell(row,colData+4+ cell.column).value == None:
                                    wsData.cell(row,colData+4+ cell.column).value =  1  #increment vote
                                else:
                                    wsData.cell(row,colData+4+ cell.column).value = wsData.cell(row,colData+4+ cell.column).value + 1

    wsData.cell(1, colData+3).value = 'Function+Classification'
    wsData.cell(1, + colData+4).value = 'Classification'

    # classification columns 
    wsData.cell(1, + colData+5).value = 'Sensor Fcns'
    wsData.cell(1, + colData+6).value = 'Control Fcns'
    wsData.cell(1, + colData+7).value = 'Actuator Fcns'
    wsData.cell(1, + colData+8).value = 'Log & Diag'
    wsData.cell(1, + colData+9).value = 'Initialization'
    wsData.cell(1, + colData+10).value = 'Calibration'
    wsData.cell(1, + colData+11).value = 'Communication'

    wbData.save(filename = dataFile)
    wbData.close()
    # print(len(hitsArray))

# analyzeData(dataFile)