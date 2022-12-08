
import csv
import openpyxl

from ExtractComments import getComments
from openpyxl import load_workbook
from functiondefextractor import core_extractor

def convert2Excel(dataFile):
    #dataFile = 'ExtractFunctions1.csv'

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(dataFile + '.csv', newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(dataFile + '.xlsx')
    wb.close()

def createDataFile(srcPath, genDataFile, dataFile):
    
    if genDataFile:
        output0 = core_extractor.extractor (srcPath, exclude=r'*.h, *.txt, wscript, Makefile.waf')
        output0.to_csv(dataFile +'.csv')  # local folder file , adds sr no #

        convert2Excel(dataFile)

    dataFile = dataFile + '.xlsx'

    wb = load_workbook(filename = dataFile)

    # grab the active worksheet
    ws = wb.active
    row_count = ws.max_row

    column = 2
    for row in range(2, row_count):
        print(row)
        getComments(row, column, ws)

    wb.save(filename = dataFile)
    wb.close()

