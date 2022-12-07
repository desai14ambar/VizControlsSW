from openpyxl import load_workbook

# dataFile = 'ExtractFunctions1'
# dataFile = dataFile + '.xlsx'

def finalVotefcn(dataFile):
    wbData = load_workbook(filename = dataFile)

    # grab the active worksheet
    wsData = wbData.active
    row_countData = wsData.max_row
    
    colData = 9

    for row in range(2, row_countData):
        voteArr = []
        for i in range(0,7):
            getCell = wsData.cell(row, colData+i)
            vote = getCell.value
            if vote == None:
                vote = 0
            voteArr.append(vote)
        maxVote = max(voteArr)
        if maxVote is not 0:
            voteIndex = voteArr.index(maxVote)
            classify = wsData.cell(1, colData + voteIndex).value
            print(row, '  ' , voteArr , ' ---> ' , max(voteArr) , '  ', voteArr.count(max(voteArr)) , '  ', voteArr.index(max(voteArr)), ' ', classify)
            wsData.cell(row, colData + 8).value = classify
        else:
            wsData.cell(row, colData + 8).value = "Unclassified"

    wbData.save(filename = dataFile)
    wbData.close()

